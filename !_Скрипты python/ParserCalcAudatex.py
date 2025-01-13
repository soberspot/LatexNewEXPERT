import PyPDF2
from openpyxl import Workbook
from tkinter import Tk, filedialog, messagebox
import re


def parse_pdf(pdf_path):
    """Чтение текста из PDF файла."""
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ''
            for page in reader.pages:
                text += page.extract_text()
        return text
    except Exception as e:
        messagebox.showerror("Error", f"Ошибка чтения PDF файла: {e}")
        return None


def extract_section_data(text, start_marker, header_marker, column_count):
    """
    Извлекает данные после заданного маркера и заголовка.
    """
    try:
        # Находим раздел после маркера
        if start_marker not in text:
            return []  # Если маркер отсутствует, возвращаем пустой список
        section = text.split(start_marker, 1)[-1]
        lines = section.split("\n")

        # Ищем строку заголовка
        header_index = next((i for i, line in enumerate(lines) if header_marker in line), None)
        if header_index is None:
            return []  # Если заголовок отсутствует, возвращаем пустой список

        # Берем строки, следующие за заголовком
        data_lines = lines[header_index + 1:]

        # Парсим строки
        parsed_data = []
        for line in data_lines:
            line = line.strip()
            if not line:  # Пропускаем пустые строки
                continue

            # Для раздела "О К Р А С К А" особая обработка
            if start_marker == "О К Р А С К А":
                match = re.match(r'^(.+?)\s{2,}(.+?)\s{2,}([\d\w\-]+)\s+([\d\.]+)$', line)
                if match:
                    parsed_data.append(list(match.groups()))
                    continue

            # Общая обработка для других разделов
            columns = re.split(r'\s{2,}', line)  # Разделяем по двум или более пробелам
            if len(columns) < column_count:
                # Если колонок меньше, добавляем пустые значения
                columns.extend([""] * (column_count - len(columns)))
            parsed_data.append(columns[:column_count])  # Ограничиваем количество колонок

        return parsed_data
    except Exception as e:
        messagebox.showerror("Error", f"Ошибка обработки данных: {e}")
        return []


def save_to_excel(parts_data, works_data, paint_data, excel_path):
    """
    Сохраняет данные в Excel файл.
    """
    try:
        wb = Workbook()
        ws = wb.active

        # Сохраняем данные раздела "Запчасти"
        ws.title = "Запчасти"
        ws.append(["УПР №", "НАЗВАНИЕ", "№ ДЕТАЛИ", "СТОИМ"])
        for row in parts_data:
            ws.append(row)

        # Добавляем лист для раздела "Работы"
        ws_works = wb.create_sheet("Работы")
        ws_works.append(["№ РАБ.", "ПОЗ./РАБОТЫ", "РП", "СТОИМ"])
        for row in works_data:
            ws_works.append(row)

        # Добавляем лист для раздела "Окраска"
        ws_paint = wb.create_sheet("Окраска")
        ws_paint.append(["ОПЕРАЦИЯ", "ОПИСАНИЕ РП", "КОД ОПЕР.", "СТОИМ"])
        for row in paint_data:
            ws_paint.append(row)

        # Сохраняем Excel
        wb.save(excel_path)
        messagebox.showinfo("Success", f"Данные сохранены в {excel_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Ошибка сохранения Excel файла: {e}")


def main():
    root = Tk()
    root.withdraw()  # Скрыть главное окно tkinter

    # Выбор PDF-файла
    pdf_path = filedialog.askopenfilename(
        title="Выберите PDF файл",
        filetypes=[("PDF files", "*.pdf")]
    )
    if not pdf_path:
        messagebox.showwarning("Warning", "Вы не выбрали PDF файл.")
        return

    # Чтение PDF
    text = parse_pdf(pdf_path)
    if not text:
        return

    # Извлечение данных для каждого раздела
    parts_data = extract_section_data(text, "З А П Ч А С Т И", "УПР №", 4)
    works_data = extract_section_data(
        text,
        "№ РАБ. ПОЗ./ РАБОТЫ ПО РЕМ./ОТДЕЛЬНЫЕ/КОМБИН. РАБОТЫ",
        "№ РАБ.",
        4
    )
    paint_data = extract_section_data(
        text,
        "О К Р А С К А",
        "ОПЕРАЦИЯ / ОПИСАНИЕ РП",
        4
    )

    # Если нет данных
    if not parts_data and not works_data and not paint_data:
        messagebox.showwarning("Warning", "В PDF файле не найдено данных.")
        return

    # Указание пути для сохранения Excel
    excel_path = filedialog.asksaveasfilename(
        title="Сохранить файл как",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not excel_path:
        messagebox.showwarning("Warning", "Вы не выбрали место для сохранения.")
        return

    # Сохранение в Excel
    save_to_excel(parts_data, works_data, paint_data, excel_path)


if __name__ == "__main__":
    main()
